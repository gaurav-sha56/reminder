#This is a reminder app this app will give you reminders for your things
#keep learning


import openpyxl
import time
from datetime import datetime
from plyer import notification



if __name__=="__main__":
    task={}
    times=[]

    path = r"C:\Users\Gaurav Sharma\Documents\Project2\reminders.xlsx"
    wb_obj = openpyxl.load_workbook(path)  #object to load workbook
    sheet_obj = wb_obj.active  #object to activae workbook data of the cell row=1 and col=1 
    row=sheet_obj.max_row
    col=sheet_obj.max_column


    for i in range(2,row+1):   # list of tasks 
        cell_obj = sheet_obj.cell(row = i, column = 1)  # taska to be performed
        cell_obj_time = sheet_obj.cell(row = i, column = 2)
        formatted_time = cell_obj_time.value.strftime("%H:%M:%S")  # Format the time as HH:MM:SS)
        task[cell_obj.value]=formatted_time
        times.append(formatted_time)
    
    

    condition = True
    while condition:
        c=datetime.now()
        current_time = c.strftime('%H:%M:%S')
        for val in task:
            if task[val]==current_time:

                notification.notify(
                    title = "Reminder",
                    message = val,
                    timeout = 2 #displaying time
                )

                time.sleep(5)
                times.remove(task[val])
        if times==[]:
            condition = False
